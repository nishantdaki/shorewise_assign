from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import JSONParser
from django.http.response import JsonResponse
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os

from djangoauth.settings import MEDIA_ROOT
from django.core.files.storage import default_storage


class SimpleApI(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request, **kwargs):
        try:
            workbook_name = "data.xlsx"
            excel_file_path = '%s/%s' % (MEDIA_ROOT, workbook_name)
            
            if os.path.exists(excel_file_path):
                
                wb = load_workbook(excel_file_path)
                page = wb.active
                max = page.max_row
                page.cell(row=max+1, column=1, value=request.data['name'])
                page.cell(row=max+1, column=2, value=request.data['dob'])
                page.cell(row=max+1, column=3, value=request.data['country'])
                wb.save(excel_file_path)
                return JsonResponse({"status":"successful"})
            return JsonResponse({"status":"failed"})
        except Exception as e:
            return JsonResponse({"status":"failed"})